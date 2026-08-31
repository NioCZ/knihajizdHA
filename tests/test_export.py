"""Tests for the executor-safe Excel exporter."""

from __future__ import annotations

import importlib.util
from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook

ROOT = Path(__file__).parents[1]
SPEC = importlib.util.spec_from_file_location(
    "kniha_jizd_export", ROOT / "custom_components/kniha_jizd/export.py"
)
assert SPEC is not None and SPEC.loader is not None
EXPORT_MODULE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(EXPORT_MODULE)


class ExportExcelTest(unittest.TestCase):
    """Verify workbook structure and representative aggregations."""

    def test_two_sheet_export_and_daily_totals(self) -> None:
        """Export a representative day and inspect typed workbook values."""
        test_output = ROOT / "test-output"
        test_output.mkdir(exist_ok=True)
        with tempfile.TemporaryDirectory(dir=test_output) as temporary_directory:
            output_path = Path(temporary_directory) / "kniha_jizd.xlsx"
            result = EXPORT_MODULE.export_excel(
                ROOT / "tests/fixtures/raw_sample.json", output_path, "2026-08"
            )

            self.assertEqual(result["month"], "2026-08")
            self.assertEqual(result["days"], 1)
            self.assertEqual(result["segments"], 3)
            workbook = load_workbook(output_path, data_only=False)
            self.assertEqual(workbook.sheetnames, ["Kniha jízd", "Raw data"])

            summary = workbook["Kniha jízd"]
            self.assertEqual(
                [cell.value for cell in summary[1]], EXPORT_MODULE.SUMMARY_COLUMNS
            )
            self.assertEqual(summary["F2"].value, 33)
            self.assertEqual(summary["G2"].value, 9)
            self.assertIn("Průmyslová 12", summary["C2"].value)
            self.assertNotIn("Vinohradská 50", summary["C2"].value)

            raw = workbook["Raw data"]
            headers = [cell.value for cell in raw[1]]
            self.assertEqual(raw.max_row, 4)
            self.assertIn("odometer_wait_timed_out", headers)
            self.assertIn("map_candidates", headers)
            self.assertIn("candidate_search_radius_m", headers)

    def test_private_only_day_hides_route_and_customer(self) -> None:
        """A private report row contains only its date and private kilometres."""
        rows = EXPORT_MODULE._build_summary_rows(
            [
                {
                    "date": "2026-08-20",
                    "started_at": "2026-08-20T08:00:00+00:00",
                    "start_address": "Soukromý domov",
                    "end_address": "Soukromý cíl",
                    "purpose": "Soukromá",
                    "trip_type": "private",
                    "distance_km": 12.5,
                }
            ]
        )

        self.assertEqual(rows[0]["Start/Odkud"], "")
        self.assertEqual(rows[0]["Přes"], "")
        self.assertEqual(rows[0]["Cíl/Kam"], "")
        self.assertEqual(rows[0]["Zákazník"], "")
        self.assertEqual(rows[0]["Soukromé km"], 13)

    def test_transient_stop_is_hidden_from_route_but_keeps_its_kilometres(self) -> None:
        """A petrol stop must not appear in Přes or as a separate customer."""
        rows = EXPORT_MODULE._build_summary_rows(
            [
                {
                    "date": "2026-08-25",
                    "started_at": "2026-08-25T08:00:00+00:00",
                    "start_address": "Kroměříž",
                    "end_address": "Benzina D1",
                    "purpose": "Nemocnice",
                    "trip_type": "business",
                    "distance_km": 40,
                    "journey_id": "journey",
                    "journey_role": "transient_stop",
                },
                {
                    "date": "2026-08-25",
                    "started_at": "2026-08-25T08:15:00+00:00",
                    "start_address": "Benzina D1",
                    "end_address": "Nemocnice Brno",
                    "purpose": "Nemocnice",
                    "trip_type": "business",
                    "distance_km": 30,
                    "journey_id": "journey",
                    "journey_role": "destination",
                },
            ]
        )

        self.assertEqual(rows[0]["Start/Odkud"], "Kroměříž")
        self.assertEqual(rows[0]["Přes"], "")
        self.assertEqual(rows[0]["Cíl/Kam"], "Nemocnice Brno")
        self.assertEqual(rows[0]["Zákazník"], "Nemocnice")
        self.assertEqual(rows[0]["Služební km"], 70)

    def test_untagged_three_minute_stop_is_also_hidden(self) -> None:
        """Historical quick stops stay out of Přes even without journey metadata."""
        rows = EXPORT_MODULE._build_summary_rows(
            [
                {
                    "date": "2026-08-25",
                    "started_at": "2026-08-25T07:44:23+00:00",
                    "ended_at": "2026-08-25T08:10:02+00:00",
                    "start_address": "Výstavní 378/18",
                    "end_address": "Masná 458/106",
                    "purpose": "",
                    "trip_type": "business",
                    "distance_km": 12,
                },
                {
                    "date": "2026-08-25",
                    "started_at": "2026-08-25T08:10:50+00:00",
                    "ended_at": "2026-08-25T08:29:18+00:00",
                    "start_address": "Masná 458/106",
                    "end_address": "Jihlavská 24",
                    "purpose": "Fakultní nemocnice",
                    "trip_type": "business",
                    "distance_km": 6,
                },
            ]
        )

        self.assertEqual(rows[0]["Start/Odkud"], "Výstavní 378/18")
        self.assertEqual(rows[0]["Přes"], "")
        self.assertEqual(rows[0]["Cíl/Kam"], "Jihlavská 24")
        self.assertEqual(rows[0]["Zákazník"], "Fakultní nemocnice")
        self.assertEqual(rows[0]["Služební km"], 18)

    def test_today_fuel_and_quick_handoff_stops_are_hidden(self) -> None:
        """Hide the observed Mořice and Novovysočanská stops but keep all km."""
        rows = EXPORT_MODULE._build_summary_rows(
            [
                {
                    "date": "2026-08-31",
                    "started_at": "2026-08-31T05:04:33+00:00",
                    "ended_at": "2026-08-31T05:22:06+00:00",
                    "start_address": "Vrchlického 699, Kroměříž",
                    "end_address": "Mořice 192, Mořice",
                    "end_latitude": 49.31964,
                    "end_longitude": 17.19965,
                    "purpose": "",
                    "trip_type": "business",
                    "classification_source": "manual_panel",
                    "journey_role": "destination",
                    "distance_km": 17,
                },
                {
                    "date": "2026-08-31",
                    "started_at": "2026-08-31T05:25:37+00:00",
                    "ended_at": "2026-08-31T08:02:59+00:00",
                    "start_address": "Mořice 194, Mořice",
                    "start_latitude": 49.31965,
                    "start_longitude": 17.19966,
                    "end_address": "Vídeňská 817/7, Praha 4-Krč",
                    "end_latitude": 50.02552,
                    "end_longitude": 14.46027,
                    "purpose": "Thomayerova nemocnice",
                    "trip_type": "business",
                    "classification_source": "learned_place",
                    "journey_role": "destination",
                    "matched_place_role": "client",
                    "distance_km": 244,
                },
                {
                    "date": "2026-08-31",
                    "started_at": "2026-08-31T14:19:40+00:00",
                    "ended_at": "2026-08-31T15:11:28+00:00",
                    "start_address": "Vídeňská 817/7, Praha 4-Krč",
                    "end_address": "Novovysočanská 976/33, Praha 9",
                    "end_latitude": 50.10087,
                    "end_longitude": 14.49290,
                    "purpose": "Thomayerova nemocnice",
                    "trip_type": "business",
                    "classification_source": "notification_return",
                    "journey_role": "return",
                    "return_of_segment_id": "thomayer",
                    "distance_km": 14,
                },
                {
                    "date": "2026-08-31",
                    "started_at": "2026-08-31T15:16:40+00:00",
                    "ended_at": "2026-08-31T15:28:24+00:00",
                    "start_address": "Novovysočanská 555/24, Praha 9",
                    "start_latitude": 50.10082,
                    "start_longitude": 14.49275,
                    "end_address": "Na Jetelce 69, Praha 9",
                    "purpose": "Altium",
                    "trip_type": "business",
                    "classification_source": "configured_company",
                    "journey_role": "destination",
                    "distance_km": 3,
                },
                {
                    "date": "2026-08-31",
                    "started_at": "2026-08-31T16:01:18+00:00",
                    "ended_at": "2026-08-31T16:25:54+00:00",
                    "start_address": "Na Jetelce 69, Praha 9",
                    "end_address": "U Slavie 1540/2a, Praha 10-Vršovice",
                    "purpose": "",
                    "trip_type": "business",
                    "classification_source": "manual_panel",
                    "journey_role": "destination",
                    "distance_km": 8,
                },
            ]
        )

        self.assertEqual(rows[0]["Start/Odkud"], "Vrchlického 699, Kroměříž")
        self.assertEqual(
            rows[0]["Přes"],
            "Vídeňská 817/7, Praha 4-Krč → Na Jetelce 69, Praha 9",
        )
        self.assertEqual(
            rows[0]["Cíl/Kam"], "U Slavie 1540/2a, Praha 10-Vršovice"
        )
        self.assertEqual(rows[0]["Zákazník"], "Thomayerova nemocnice, Altium")
        self.assertEqual(rows[0]["Služební km"], 286)
        self.assertNotIn("Mořice", rows[0]["Přes"])
        self.assertNotIn("Novovysočanská", rows[0]["Přes"])

    def test_short_confirmed_client_visit_remains_visible(self) -> None:
        """Do not hide a genuine client merely because the visit was brief."""
        rows = EXPORT_MODULE._build_summary_rows(
            [
                {
                    "date": "2026-08-31",
                    "started_at": "2026-08-31T08:00:00+00:00",
                    "ended_at": "2026-08-31T08:30:00+00:00",
                    "start_address": "Firma",
                    "end_address": "Laboratoř",
                    "end_latitude": 50.0,
                    "end_longitude": 14.0,
                    "purpose": "Potvrzený klient",
                    "trip_type": "business",
                    "classification_source": "learned_place",
                    "journey_role": "destination",
                    "matched_place_role": "client",
                    "distance_km": 20,
                },
                {
                    "date": "2026-08-31",
                    "started_at": "2026-08-31T08:36:00+00:00",
                    "ended_at": "2026-08-31T09:00:00+00:00",
                    "start_address": "Vedlejší vchod laboratoře",
                    "start_latitude": 50.0001,
                    "start_longitude": 14.0001,
                    "end_address": "Firma",
                    "purpose": "",
                    "trip_type": "business",
                    "classification_source": "manual_panel",
                    "journey_role": "return",
                    "distance_km": 20,
                },
            ]
        )

        self.assertEqual(rows[0]["Přes"], "Laboratoř")
        self.assertEqual(rows[0]["Zákazník"], "Potvrzený klient")
        self.assertEqual(rows[0]["Služební km"], 40)

    def test_summary_uses_exact_configured_home_and_company_addresses(self) -> None:
        """Nearby mobile fixes are canonicalized only in the daily summary."""
        segments = [
            {
                "date": "2026-08-25",
                "started_at": "2026-08-25T06:10:00+00:00",
                "start_address": "Vrchlického 682/4, Kroměříž",
                "start_latitude": 49.29810,
                "start_longitude": 17.39210,
                "end_address": "Výstavní 380/20, Brno",
                "end_latitude": 49.18920,
                "end_longitude": 16.58420,
                "purpose": "Firma",
                "trip_type": "business",
                "distance_km": 66,
                "configured_place": "company",
            },
            {
                "date": "2026-08-25",
                "started_at": "2026-08-25T14:00:00+00:00",
                "start_address": "Výstavní 380/20, Brno",
                "start_latitude": 49.18920,
                "start_longitude": 16.58420,
                "end_address": "Vrchlického 682/4, Kroměříž",
                "end_latitude": 49.29810,
                "end_longitude": 17.39210,
                "purpose": "Návrat",
                "trip_type": "business",
                "distance_km": 66,
                "configured_place": "home",
            },
        ]
        configured_places = {
            "home": {
                "address": "Vrchlického 699/2, Kroměříž",
                "latitude": 49.29800,
                "longitude": 17.39200,
                "radius_m": 300,
            },
            "company": {
                "address": "Výstavní 378/18, Brno",
                "latitude": 49.18910,
                "longitude": 16.58410,
                "radius_m": 300,
            },
        }

        rows = EXPORT_MODULE._build_summary_rows(segments, configured_places)

        self.assertEqual(rows[0]["Start/Odkud"], "Vrchlického 699/2, Kroměříž")
        self.assertEqual(rows[0]["Přes"], "Výstavní 378/18, Brno")
        self.assertEqual(rows[0]["Cíl/Kam"], "Vrchlického 699/2, Kroměříž")
        self.assertEqual(segments[0]["start_address"], "Vrchlického 682/4, Kroměříž")
        self.assertEqual(segments[1]["end_address"], "Vrchlického 682/4, Kroměříž")

    def test_month_filter_excludes_other_periods(self) -> None:
        """Keep both workbook sheets inside the requested calendar month."""
        test_output = ROOT / "test-output"
        test_output.mkdir(exist_ok=True)
        with tempfile.TemporaryDirectory(dir=test_output) as temporary_directory:
            output_path = Path(temporary_directory) / "kniha_jizd_2026-07.xlsx"
            result = EXPORT_MODULE.export_excel(
                ROOT / "tests/fixtures/raw_sample.json", output_path, "2026-07"
            )

            self.assertEqual(result["days"], 0)
            self.assertEqual(result["segments"], 0)
            workbook = load_workbook(output_path, data_only=False)
            self.assertEqual(workbook["Kniha jízd"].max_row, 1)
            self.assertEqual(workbook["Raw data"].max_row, 1)

    def test_month_filter_rejects_invalid_value(self) -> None:
        """Reject ambiguous dates before creating a misleading report."""
        with self.assertRaisesRegex(ValueError, "YYYY-MM"):
            EXPORT_MODULE.export_excel(
                ROOT / "tests/fixtures/raw_sample.json",
                ROOT / "test-output/invalid.xlsx",
                "08/2026",
            )


if __name__ == "__main__":
    unittest.main()
