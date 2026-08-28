"""Pandas Excel export for Kniha jízd."""

from __future__ import annotations

from datetime import UTC, datetime
import json
from math import asin, cos, floor, radians, sin, sqrt
from pathlib import Path
import re
from typing import Any


SUMMARY_COLUMNS = [
    "Datum",
    "Start/Odkud",
    "Přes",
    "Cíl/Kam",
    "Zákazník",
    "Služební km",
    "Soukromé km",
]
_IMPLICIT_TRANSIENT_STOP_SECONDS = 3 * 60


def export_excel(
    raw_path: Path,
    output_path: Path,
    month: str | None = None,
    configured_places: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Build a two-sheet xlsx file, optionally restricted to one month."""
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with raw_path.open("r", encoding="utf-8") as file_handle:
        raw_document = json.load(file_handle)
    segments = raw_document.get("segments", [])
    if not isinstance(segments, list):
        raise ValueError("Raw data file does not contain a 'segments' list")
    if month is not None:
        if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", month):
            raise ValueError("Month must use the YYYY-MM format")
        segments = [
            segment
            for segment in segments
            if isinstance(segment, dict) and _segment_belongs_to_month(segment, month)
        ]

    summary_rows = _build_summary_rows(segments, configured_places)
    summary_frame = pd.DataFrame(summary_rows, columns=SUMMARY_COLUMNS)
    if not summary_frame.empty:
        summary_frame["Datum"] = pd.to_datetime(
            summary_frame["Datum"], errors="coerce"
        ).dt.date
    raw_frame = pd.DataFrame.from_records(segments)
    if raw_frame.empty:
        raw_frame = pd.DataFrame(columns=_raw_columns())

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_frame.to_excel(writer, sheet_name="Kniha jízd", index=False)
        raw_frame.to_excel(writer, sheet_name="Raw data", index=False)

        for worksheet in writer.book.worksheets:
            worksheet.sheet_view.showGridLines = False
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.row_dimensions[1].height = 24
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for column_cells in worksheet.columns:
                max_length = max(
                    (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                    default=0,
                )
                column_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[column_letter].width = min(
                    max(max_length + 2, 11), 65
                )
                for cell in column_cells[1:]:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        summary_sheet = writer.book["Kniha jízd"]
        for row in range(2, summary_sheet.max_row + 1):
            summary_sheet.cell(row=row, column=1).number_format = "yyyy-mm-dd"
            summary_sheet.cell(row=row, column=6).number_format = "0"
            summary_sheet.cell(row=row, column=7).number_format = "0"

    return {
        "path": str(output_path),
        "month": month,
        "days": len(summary_rows),
        "segments": len(segments),
    }


def _segment_belongs_to_month(segment: dict[str, Any], month: str) -> bool:
    """Return whether the segment's local date belongs to the selected month."""
    date = str(segment.get("date") or _date_from_timestamp(segment.get("started_at")))
    return date.startswith(f"{month}-")


def _build_summary_rows(
    segments: list[dict[str, Any]],
    configured_places: dict[str, dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Aggregate raw segments by local calendar date."""
    by_date: dict[str, list[dict[str, Any]]] = {}
    for segment in segments:
        date = str(segment.get("date") or _date_from_timestamp(segment.get("started_at")))
        by_date.setdefault(date, []).append(segment)

    rows: list[dict[str, Any]] = []
    for date in sorted(by_date):
        day_segments = sorted(
            by_date[date], key=lambda item: str(item.get("started_at", ""))
        )
        business_segments = [
            segment
            for segment in day_segments
            if segment.get("trip_type") == "business"
        ]
        visible_business_segments = []
        for index, segment in enumerate(business_segments):
            next_segment = (
                business_segments[index + 1]
                if index + 1 < len(business_segments)
                else None
            )
            if segment.get("journey_role") == "transient_stop":
                continue
            if next_segment is not None and _is_very_short_stop(
                segment, next_segment
            ):
                continue
            visible_business_segments.append(segment)
        route_nodes: list[str] = []
        if business_segments:
            route_nodes.append(
                _summary_route_address(
                    business_segments[0], "start", configured_places
                )
            )
            route_nodes.extend(
                _summary_route_address(segment, "end", configured_places)
                for segment in visible_business_segments
            )
        route_nodes = _deduplicate_adjacent(route_nodes)

        customers = _unique_nonempty(
            str(segment.get("purpose") or "")
            for segment in visible_business_segments
        )
        business_km = sum(
            _number(segment.get("distance_km"))
            for segment in business_segments
        )
        private_km = sum(
            _number(segment.get("distance_km"))
            for segment in day_segments
            if segment.get("trip_type") == "private"
        )

        rows.append(
            {
                "Datum": date,
                "Start/Odkud": route_nodes[0] if route_nodes else "",
                "Přes": " → ".join(route_nodes[1:-1]),
                "Cíl/Kam": route_nodes[-1] if route_nodes else "",
                "Zákazník": ", ".join(customers),
                "Služební km": _whole_km(business_km),
                "Soukromé km": _whole_km(private_km),
            }
        )
    return rows


def _summary_route_address(
    segment: dict[str, Any],
    side: str,
    configured_places: dict[str, dict[str, Any]] | None,
) -> str:
    """Use an exact configured address in the summary without changing raw data."""
    observed = str(segment.get(f"{side}_address") or "")
    if not configured_places:
        return observed

    explicit_place = segment.get(
        "start_configured_place" if side == "start" else "configured_place"
    )
    if isinstance(explicit_place, str):
        explicit_rule = configured_places.get(explicit_place)
        if isinstance(explicit_rule, dict):
            configured_address = str(explicit_rule.get("address") or "").strip()
            if configured_address:
                return configured_address

    for place_name in ("home", "company"):
        place = configured_places.get(place_name)
        if not isinstance(place, dict):
            continue
        configured_address = str(place.get("address") or "").strip()
        if configured_address and _segment_side_matches_place(segment, side, place):
            return configured_address
    return observed


def _segment_side_matches_place(
    segment: dict[str, Any], side: str, place: dict[str, Any]
) -> bool:
    """Match one stored route endpoint against a configured GPS zone."""
    distance_m = _coordinate_distance_m(
        segment.get(f"{side}_latitude"),
        segment.get(f"{side}_longitude"),
        place.get("latitude"),
        place.get("longitude"),
    )
    if distance_m is None:
        return False
    try:
        radius_m = float(place.get("radius_m"))
    except (TypeError, ValueError):
        return False
    return distance_m <= radius_m


def _coordinate_distance_m(
    latitude: Any,
    longitude: Any,
    reference_latitude: Any,
    reference_longitude: Any,
) -> float | None:
    """Return the distance between two GPS points in metres."""
    try:
        latitude_value = float(latitude)
        longitude_value = float(longitude)
        reference_latitude_value = float(reference_latitude)
        reference_longitude_value = float(reference_longitude)
    except (TypeError, ValueError):
        return None
    earth_radius_m = 6_371_000.0
    delta_latitude = radians(reference_latitude_value - latitude_value)
    delta_longitude = radians(reference_longitude_value - longitude_value)
    a = (
        sin(delta_latitude / 2) ** 2
        + cos(radians(latitude_value))
        * cos(radians(reference_latitude_value))
        * sin(delta_longitude / 2) ** 2
    )
    return 2 * earth_radius_m * asin(sqrt(a))


def _date_from_timestamp(value: Any) -> str:
    """Use the ISO date prefix only as compatibility fallback."""
    text = str(value or "")
    return text[:10] if len(text) >= 10 else "Neznámé datum"


def _number(value: Any) -> float:
    """Coerce a JSON value to a summable float."""
    try:
        return float(value) if value is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def _whole_km(value: float) -> int:
    """Round exported kilometre totals to a whole number."""
    return int(floor(max(0.0, value) + 0.5))


def _is_very_short_stop(
    segment: dict[str, Any], next_segment: dict[str, Any]
) -> bool:
    """Hide an untagged stop when the next leg leaves the same place immediately."""
    ended_at = _parse_iso_datetime(segment.get("ended_at"))
    next_started_at = _parse_iso_datetime(next_segment.get("started_at"))
    if ended_at is None or next_started_at is None:
        return False
    gap_seconds = (next_started_at - ended_at).total_seconds()
    if gap_seconds < 0 or gap_seconds > _IMPLICIT_TRANSIENT_STOP_SECONDS:
        return False
    end_address = str(segment.get("end_address") or "").strip().casefold()
    next_start_address = str(next_segment.get("start_address") or "").strip().casefold()
    return bool(end_address and end_address == next_start_address)


def _parse_iso_datetime(value: Any) -> datetime | None:
    """Parse one stored ISO timestamp for short-stop comparisons."""
    try:
        parsed = datetime.fromisoformat(str(value or "").replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    return parsed.astimezone(UTC)


def _deduplicate_adjacent(values: list[str]) -> list[str]:
    """Remove empty and immediately repeated route nodes."""
    result: list[str] = []
    for value in values:
        stripped = value.strip()
        if stripped and (not result or result[-1] != stripped):
            result.append(stripped)
    return result


def _unique_nonempty(values: Any) -> list[str]:
    """Return nonempty values in first-seen order."""
    result: list[str] = []
    for value in values:
        stripped = value.strip()
        if stripped and stripped not in result:
            result.append(stripped)
    return result


def _raw_columns() -> list[str]:
    """Keep Raw data useful even before the first recorded segment."""
    return [
        "id",
        "date",
        "started_at",
        "ended_at",
        "odometer_updated_at",
        "odometer_wait_timed_out",
        "odometer_ready",
        "odometer_shared_update",
        "odometer_completion_source",
        "odometer_rollover_at",
        "start_odometer_km",
        "end_odometer_km",
        "distance_km",
        "distance_km_raw",
        "distance_hint_km",
        "distance_reconciliation_source",
        "distance_rounding_method",
        "distance_anchor_start_km",
        "distance_anchor_end_km",
        "odometer_reconciliation_boundary_km",
        "odometer_reconciliation_boundary_source",
        "odometer_anchor_ignored_due_to_daily_conflict",
        "odometer_anchor_ignored_due_to_later_trip_start",
        "daily_odometer_override_reason",
        "daily_odometer_authoritative_total_km",
        "manual_distance_override",
        "start_address",
        "end_address",
        "start_address_raw",
        "end_address_raw",
        "start_address_manual",
        "end_address_manual",
        "start_latitude",
        "start_longitude",
        "end_latitude",
        "end_longitude",
        "purpose",
        "trip_type",
        "classification_source",
        "classification_prepared",
        "classification_ready",
        "classification_options",
        "persisted",
        "manually_edited_at",
        "notification_sent_at",
        "notification_tag",
        "journey_id",
        "journey_segment_count",
        "journey_distance_km",
        "journey_distance_complete",
        "journey_role",
        "journey_inherited_from_segment_id",
        "transient_stop",
        "transient_continuation",
        "continued_by_segment_id",
        "continuation",
        "return_of_segment_id",
        "return_context",
        "matched_place_id",
        "return_destination_label",
        "map_estimate",
        "map_address",
        "map_attribution",
        "map_candidates",
        "candidate_search_radius_m",
        "selected_map_candidate",
        "configured_place",
        "configured_place_match",
        "end_location_ready",
        "end_location_source",
        "end_location_captured_at",
        "end_location_initial_address",
        "end_location_initial_address_raw",
        "end_location_initial_latitude",
        "end_location_initial_longitude",
        "location_update_requested_at",
        "validation_error",
    ]
